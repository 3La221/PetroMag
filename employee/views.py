import os
from django.http import FileResponse, HttpResponse
from django.shortcuts import render,redirect
import json 
from openpyxl import load_workbook
import shutil
from .models import Employee,ChefStation
from datetime import datetime
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from django.template import loader


def home(request):
    
    if not request.user.is_authenticated:
        return redirect('login')
    current_month = datetime.now().month
    today = datetime.now().day 
    default_excel = "./excel/default.xlsx"
    data = []
    cells = []
    chef_station = ChefStation.objects.get(user=request.user)
    employees = Employee.objects.filter(station = chef_station.station)

    
    if (not employees):
        return render(request,'main/home.html', {'data':data,'days':[1,2,3,4,5,6,7,8,9,10]})
    for employee in employees:
        tmp = {}
        path=f"./excel/{employee.matricule}2024.xlsx"
        # print(path)
        try:
            worbook = load_workbook(path)
            sheet = worbook.active 
        except:
            shutil.copy(default_excel,path)
            worbook = load_workbook(path)
            sheet = worbook.active 
            sheet['U6'] = employee.matricule
            sheet['B6'] = employee.nom 
            sheet['B7'] = employee.prenom 
            sheet['B8'] = employee.fonction
            sheet['AG6'] = employee.date_recrut
            sheet['AG8'] = employee.date_detach
            sheet['AG9'] = employee.affect_origine
            sheet['AG10'] = employee.sit_fam
            sheet['AG11'] = employee.nbr_enfant
            worbook.save(f"./excel/{employee.matricule}2024.xlsx")
        tmp['employee'] = employee
        tmp['results'] = []
        
        if today <= 10:
            for i in range(ord('B'), ord('K')+1):
                cell_value = sheet[f'{chr(i)}{str(current_month+13)}'].value
                tmp['results'].append(cell_value)
                cells.append(f'{chr(i)}{str(current_month+13)}')
            days = [1,2,3,4,5,6,7,8,9,10]
            
        elif today <=20 :
            for i in range(ord('L'), ord('U')+1):
                cell_value = sheet[f'{chr(i)}{str(current_month+13)}'].value
                tmp['results'].append(cell_value)
                cells.append(f'{chr(i)}{str(current_month+13)}')               
            days = [11,12,13,14,15,16,17,18,19,20]
            
        elif today <=31:
            for i in range(ord('V'), ord('Z')+1):
                cell_value = sheet[f'{chr(i)}{str(current_month+13)}'].value
                tmp['results'].append(cell_value)
                cells.append(f'{chr(i)}{str(current_month+13)}')               
                
            for i in range(ord('A') , ord('D')+1):
                cell_value = sheet[f'A{chr(i)}{str(current_month+13)}'].value
                tmp['results'].append(cell_value)
                days = [21,22,23,24,25,26,27,28,29]
                cells.append(f'A{chr(i)}{str(current_month+13)}') 
            if current_month != 2 :
                cell_value = sheet[f'AE{str(current_month+13)}'].value
                tmp['results'].append(cell_value)
                cells.append(f'AE{chr(i)}{str(current_month+13)}') 
                days.append(30)
                if current_month not in [4,6,9,11]:
                    cell_value = sheet[f'AF{str(current_month+13)}'].value
                    cells.append(f'AF{chr(i)}{str(current_month+13)}') 
                    tmp['results'].append(cell_value)
                    days.append(31)         
        data.append(tmp)
    worbook.close()
    
    if request.method == "POST":
        data = json.loads(request.body.decode('utf-8'))
        worbook = load_workbook(path)
        sheet = worbook.active 
        print(data)
        for matricule,value in data.items() :
            print(matricule)
            path=f"./excel/{matricule}2024.xlsx"
            worbook = load_workbook(path)
            sheet = worbook.active 
            inner_dict = data[matricule] 
            for index,value in inner_dict.items():
                if value != '-':
                    print(cells[int(index)-1])
                    sheet[cells[int(index)-1]] = value
            print("MAT = ",matricule)
            worbook.save(f"./excel/{matricule}2024.xlsx")    
        worbook.close()
        
        return render(request,'main/home.html',{'data':data,'days':days , 'today':today, 'chef':chef_station})
    
    
    return render(request,'main/home.html',{'data':data,'days':days , 'today':today , 'chef':chef_station})


def ajout_obv(request, matricule):
    path = f"./excel/{matricule}2024.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active 

    if request.method == "POST":
        data = json.loads(request.body.decode('utf-8'))
        obv = data['observation_text']
        date = data['observation_date']
        # Split the date string and take only day and month
        _ , month, day = date.split('-')
        # Format the date as "day-month"
        day_month = f"{day}-{month}"
        i = 29 
        while True:
            cell_value = sheet[f'V{i}'].value if sheet.cell(row=i, column=22).value else None
            if cell_value is None:
                break
            i += 1
        sheet[f'V{i}'] = day_month
        sheet[f'Y{i}'] = obv        
        workbook.save(path)
        workbook.close()
        return redirect("home")
    
def download_excel(request, matricule):
    path = f"./excel/{matricule}2024.xlsx"
    if os.path.exists(path):
        response = FileResponse(open(path, 'rb'))
        response['Content-Disposition'] = 'attachment; filename="{}2024.xlsx"'.format(matricule)
        return response
    else:
        # Handle the case where the file is not found
        return HttpResponse("The file you are trying to download does not exist.", status=404)



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # Check if the authenticated user is a chef station
            chef_station = ChefStation.objects.filter(user=user).first()
            if chef_station is not None:
                login(request, user)
                # Redirect to the dashboard or desired page
                return redirect('home')  # Replace 'dashboard' with your URL name
            else:
                messages.error(request, 'You are not authorized to access this page.')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'main/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')





